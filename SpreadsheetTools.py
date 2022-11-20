
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl import load_workbook


def delete_all_rows(ws: Worksheet) -> None:
    '''
    Deletes all the rows in a worksheet.
    '''
    ws.delete_rows(1, ws.max_row)


def delete_table(workbook: Workbook, table_name: str) -> Worksheet:
    '''
    Deletes an existing table from a workbook and returns the
    worksheet that previously contained the table.
    '''
    for ws in workbook.worksheets:
        for table in ws.tables.values():
            if table.name == table_name:
                for row in ws[table.ref]:
                    for cell in row:
                        cell.value = None
                del ws.tables[table_name]
                return ws
    return None


def new_table(ws: Worksheet, table_headers: list[str], table_rows: list[list | dict], table_name: str) -> None:

    # Be sure to first delete the existing rows if you want the table to start on row 1
    delete_all_rows(ws)

    ws.append(table_headers)

    # add rows of data
    for row in table_rows:
        if isinstance(row, dict):
            ws.append([row[h] for h in table_headers])
        else:
            ws.append(row)

    xlscols = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z']
    for i in range(26):
        for j in range(26):
            xlscols.append(xlscols[i] + xlscols[j])

    table = Table(
        displayName=table_name,
        ref=f'A1:{xlscols[len(table_headers)-1]}{len(table_rows) + 1}'
    )

    # Add a default style with striped rows and banded columns
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=True
    )

    ws.add_table(table)


def new_wb_with_table(table_headers: list[str], table_rows: list[list | dict], table_name: str, sheet_name: str = 'Sheet1') -> Workbook:
    '''
    Creates and returns a new workbook containing a new worksheet that contains the new table.
    '''
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    new_table(ws, table_headers, table_rows, table_name)
    return wb


def new_wb_with_tables(descriptors: list[dict]) -> Workbook:
    '''
    Creates and returns a new workbook containing new worksheets each containing a new table.
    The descriptors param is a list of dicts that should look like this:
    [
        {
            table_headers: list[str],
            table_rows: list[list | dict],
            table_name: str,
            sheet_name: str = 'Sheet1'
        },
        {
            table_headers: list[str],
            table_rows: list[list | dict],
            table_name: str,
            sheet_name: str = 'Sheet1'
        }
    ]
    '''
    wb = Workbook()
    active_sheet_name = wb.active.title
    for d in descriptors:
        new_table(wb.create_sheet(d['sheet_name']), d['table_headers'], d['table_rows'], d['table_name'])
    del wb[active_sheet_name]
    return wb


def replace_table_in_existing_wb(filename: str, table_headers: list[str], table_rows: list[list | dict], table_name: str) -> None:
    '''
    Opens a workbook file, deletes the current table, then replaces it with the new table and saves the file.
    '''
    wb = load_workbook(filename=filename)
    ws = delete_table(wb, table_name)
    new_table(ws, table_headers, table_rows, table_name)
    wb.save(filename)


if __name__ == '__main__':
    filename = './table_test.xlsx'
    '''data = [
        {"Fruit": 'Apples', "2011": 10000, "2012": 5000, "2013": 8000, "2014": 6000},
        {"Fruit": 'Pears', "2011": 2000, "2012": 3000, "2013": 4000, "2014": 5000},
        {"Fruit": 'Bananas', "2011": 6000, "2012": 6000, "2013": 6500, "2014": 6000},
        {"Fruit": 'Oranges', "2011": 500, "2012": 300, "2013": 200, "2014": 700000},
    ]
    head = ["Fruit", "2011", "2012", "2013", "2014"]
    wb = new_wb_with_table(head, data, 'MY_TABLE', 'MY_SHEET')
    wb.save(filename)'''

    '''
    wb = load_workbook(filename=filename)
    ws = wb['MY_SHEET']
    delete_all_rows(ws)
    wb.save(filename)
    '''

    '''data = [
        {
            'table_headers': ["Fruit", "2011", "2012", "2013", "2014"],
            'table_rows': [
                {"Fruit": 'Apples', "2011": 10000, "2012": 5000, "2013": 8000, "2014": 6000},
                {"Fruit": 'Pears', "2011": 2000, "2012": 3000, "2013": 4000, "2014": 5000},
                {"Fruit": 'Bananas', "2011": 6000, "2012": 6000, "2013": 6500, "2014": 6000},
                {"Fruit": 'Oranges', "2011": 500, "2012": 300, "2013": 200, "2014": 700000}
            ],
            'sheet_name': 'first sheet',
            'table_name': 'first_table'
        },
        {
            'table_headers': ["Fruit", "2011", "2012", "2013", "2014"],
            'table_rows': [
                {"Fruit": 'Apples', "2011": 10000, "2012": 5000, "2013": 8000, "2014": 6000},
                {"Fruit": 'Pears', "2011": 2000, "2012": 3000, "2013": 4000, "2014": 5000},
                {"Fruit": 'Bananas', "2011": 6000, "2012": 6000, "2013": 6500, "2014": 6000},
                {"Fruit": 'Oranges', "2011": 500, "2012": 300, "2013": 200, "2014": 700000}
            ],
            'sheet_name': 'second sheet',
            'table_name': 'second_table'
        }
    ]

    wb = new_wb_with_tables(data)
    wb.save(filename)'''

    table_headers = ["Fruit", "2011", "2012", "2013", "2014"]
    table_rows = [
        {"Fruit": 'Apples', "2011": 10000, "2012": 5000, "2013": 8000, "2014": 6000},
        {"Fruit": 'Pears', "2011": 2000, "2012": 3000, "2013": 4000, "2014": 5000},
        {"Fruit": 'Bananas', "2011": 6000, "2012": 6000, "2013": 6500, "2014": 6000},
        {"Fruit": 'Oranges', "2011": 500, "2012": 300, "2013": 200, "2014": 700000},
        {"Fruit": 'Apples', "2011": 10000, "2012": 5000, "2013": 8000, "2014": 6000},
        {"Fruit": 'Pears', "2011": 2000, "2012": 3000, "2013": 4000, "2014": 5000},
        {"Fruit": 'Bananas', "2011": 6000, "2012": 6000, "2013": 6500, "2014": 6000},
        {"Fruit": 'Oranges', "2011": 500, "2012": 300, "2013": 200, "2014": 700000},
        {"Fruit": 'Apples', "2011": 10000, "2012": 5000, "2013": 8000, "2014": 6000},
        {"Fruit": 'Pears', "2011": 2000, "2012": 3000, "2013": 4000, "2014": 5000},
        {"Fruit": 'Bananas', "2011": 6000, "2012": 6000, "2013": 6500, "2014": 6000},
        {"Fruit": 'Oranges', "2011": 500, "2012": 300, "2013": 200, "2014": 700000},
        {"Fruit": 'Apples', "2011": 10000, "2012": 5000, "2013": 8000, "2014": 6000},
        {"Fruit": 'Pears', "2011": 2000, "2012": 3000, "2013": 4000, "2014": 5000},
        {"Fruit": 'Bananas', "2011": 6000, "2012": 6000, "2013": 6500, "2014": 6000},
        {"Fruit": 'Oranges', "2011": 500, "2012": 300, "2013": 200, "2014": 700000}
    ]
    sheet_name = 'second sheet'
    table_name = 'second_table'

    replace_table_in_existing_wb(filename, table_headers, table_rows, table_name)